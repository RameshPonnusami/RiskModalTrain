import pandas as pd
import io
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Create a sample dataframe
df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6], 'C': [7, 8, 9]})

# Define the number of rows to assign strings
n = 2

# Create a list of strings to assign
strings = ["String 1", "String 2"]

# Specify the desired sheet name
sheet_name = "My Data"

# Create a list of further string rows
furt_rows = ["Further string 1", "Further string 2"]

# Create an in-memory Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.create_sheet(sheet_name)

# Write the strings to the first n rows
for i in range(n):
    worksheet.cell(row=i + 1, column=1).value = strings[i]

# Write the dataframe directly to an in-memory buffer starting from the second row
with io.BytesIO() as buffer:
    writer = pd.ExcelWriter(buffer)
    df.to_excel(writer, sheet_name=sheet_name, startrow=n + 1, header=False, index=False)
    workbook = writer.book
    print(workbook)
    worksheet = workbook.worksheets[0]


    worksheet = workbook[sheet_name]

#     part = MIMEBase("application", "octet-stream")  # Define part for attachment
#     part.set_payload(buffer.read())
#     encoders.encode_base64(part)

#     writer.close()

# Create a new in-memory workbook from the buffer
#     workbook = writer.book
#     worksheet = workbook[sheet_name]  # Access the active worksheet

    # Insert further string rows after the DataFrame
    for i, row in enumerate(furt_rows):
        worksheet.cell(row=df.shape[0] + n + 2 + i, column=1).value = row  # Adjust row index if needed
    workbook.save("my_local_file.xlsx")
# Email setup (replace with your credentials)
sender_email = "your_email@gmail.com"
receiver_email = "recipient_email@example.com"
password = "your_email_password"
#
# # Create the email message
# message = MIMEMultipart()
# message["From"] = sender_email
# message["To"] = receiver_email
# message["Subject"] = "Excel Data"
#
# # Attach the in-memory Excel buffer as an attachment
# file_name = "my_data.xlsx"
# part = MIMEBase("application", "octet-stream")
# part.set_payload(buffer.read())
# encoders.encode_base64(part)
# part.add_header(
#     "Content-Disposition",
#     f"attachment; filename={file_name}",
# )
# message.attach(part)
#
# # Send the email
# with smtplib.SMTP("smtp.gmail.com", 587) as server:
#     server.starttls()
#     server.login(sender_email, password)
#     text = message.as_string()
#     server.sendmail(sender_email, receiver_email, text)
#
# print("Email sent successfully!")
#

